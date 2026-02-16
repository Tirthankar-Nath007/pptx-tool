from openpyxl import load_workbook
from openpyxl.styles import numbers
import io
import tempfile
from pptx_backend.app import download_template

# Test that we can download the template
print("Testing template generation...")
try:
    # Call the download_template function (this will return a FileResponse)
    # We need to mock the request context, but since it's a simple function, let's call it directly
    from fastapi.testclient import TestClient
    from pptx_backend.app import app
    
    client = TestClient(app)
    
    # Download the template
    response = client.get("/download-template")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Save the template to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(response.content)
        tmp_filename = tmp.name
    
    print("Template downloaded successfully!")
    
    # Load the template
    wb = load_workbook(tmp_filename)
    ws = wb.active
    
    # Check headers
    assert ws.cell(row=2, column=1).value == "Sl no."
    assert ws.cell(row=2, column=2).value == "Brief about change (96)"
    assert ws.cell(row=2, column=3).value == "What is the impact (84)"
    assert ws.cell(row=2, column=4).value == "Dev effort"
    assert ws.cell(row=2, column=5).value == "Remarks (60)"
    assert ws.cell(row=2, column=6).value == "Gone Live/ETA"
    assert ws.cell(row=2, column=7).value == "Status"
    
    print("Headers verified!")
    
    # Check date format
    date_column = ws["F"]
    for cell in date_column[2:10]:  # Check first few cells in date column
        assert cell.number_format == numbers.FORMAT_DATE_DDMMYYYY
    print("Date format (DD/MM/YYYY) verified!")
    
    # Check data validation
    assert len(ws.data_validations.dataValidation) > 0
    print("Data validation rules found!")
    
    # Check effort options
    effort_validations = [dv for dv in ws.data_validations.dataValidation if dv.sqref.startswith("D")]
    assert len(effort_validations) == 1
    assert effort_validations[0].type == "list"
    assert "S,M,L,XL" in str(effort_validations[0].formula1)
    print("Effort options (S, M, L, XL) verified!")
    
    # Check date validation
    date_validations = [dv for dv in ws.data_validations.dataValidation if dv.sqref.startswith("F")]
    assert len(date_validations) == 1
    assert date_validations[0].type == "date"
    print("Date validation rules verified!")
    
    print("\nAll template tests passed!")
    
except Exception as e:
    print(f"\nError testing template: {str(e)}")